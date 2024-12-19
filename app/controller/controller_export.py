import time
import asyncio
import traceback
import pandas as pd
from telethon import TelegramClient
from PyQt5.QtCore import pyqtSignal, QThread
from app.utils.utils import get_display_width
from openpyxl.reader.excel import load_workbook
from PyQt5.QtWidgets import QInputDialog, QLineEdit
from telethon.errors import SessionPasswordNeededError
from openpyxl.worksheet.table import Table, TableStyleInfo


class ExportThread(QThread):
    finished_signal = pyqtSignal(bool)
    dialog_signal = pyqtSignal(str, str, bool, object)  # 定义信号

    def __init__(self, api_name, api_id, api_hash, phone_number, proxy=None, parent=None):
        super().__init__(parent)
        self.api_name = api_name
        self.api_id = api_id
        self.api_hash = api_hash
        self.phone_number = phone_number
        self.proxy = proxy
        self.input_value = None

    def run(self):
        try:
            result = asyncio.run(self.get_telegram_info(
                self.api_name, self.api_id, self.api_hash, self.phone_number, self.proxy
            ))
            self.finished_signal.emit(result)
        except Exception as e:
            print(f"线程执行出错: {e}")
            traceback.print_exc()
            self.finished_signal.emit(False)

    async def get_telegram_info(self, api_name, api_id, api_hash, phone_number, proxy=None):
        # 创建并连接客户端
        print(api_name, api_id, api_hash, phone_number, proxy)
        client = TelegramClient(api_name, api_id, api_hash, proxy=proxy)
        try:
            await client.connect()
            if not await client.is_user_authorized():
                # 使用 PyQt 的 QInputDialog 获取验证码
                send_code_response = await client.send_code_request(phone_number)
                print(send_code_response)
                # 在子线程中发送信号，请求获取密码
                self.dialog_signal.emit("验证码", "请输入验证码:", True, self)

                # 等待输入值被设置
                while self.input_value is None:
                    time.sleep(0.1)

                if self.input_value:
                    print(self.input_value)
                    verification_code = self.input_value
                    self.input_value = None
                    try:
                        await client.sign_in(phone_number, verification_code)
                    except SessionPasswordNeededError:
                        # 在子线程中发送信号，请求获取密码
                        self.dialog_signal.emit("两次验证", "请输入两次验证的密码:", True, self)

                        # 等待输入值被设置
                        while self.input_value is None:
                            time.sleep(0.1)

                        if self.input_value:
                            print(self.input_value)
                            password = self.input_value
                            self.input_value = None
                            await client.sign_in(password=password)

            # 获取对话
            groups_and_channels = []
            bots = []

            async for dialog in client.iter_dialogs():
                if dialog.is_group:
                    groups_and_channels.append((dialog.entity, '群组', dialog.name))
                elif dialog.is_channel:
                    groups_and_channels.append((dialog.entity, '频道', dialog.name))
                elif dialog.entity.bot:
                    bots.append((dialog.entity.username, dialog.entity.first_name, dialog.entity.id))

            # 准备导出数据
            group_channel_data = []
            for entity, chat_type, chat_title in groups_and_channels:
                try:
                    username = getattr(entity, 'username', None)
                    if username:
                        invite_link = f"https://t.me/{username}"
                    else:
                        invite_link = ""
                    group_channel_data.append({
                        "类型": chat_type,
                        "ID": entity.id,
                        "名称": chat_title,
                        "邀请链接": invite_link
                    })
                except Exception:
                    print(traceback.format_exc())
                    group_channel_data.append({
                        "类型": chat_type,
                        "ID": entity.id,
                        "名称": chat_title,
                        "邀请链接": ""
                    })

            bot_data = []
            for username, first_name, user_id in bots:
                bot_data.append({
                    "用户名": f"@{username}",
                    "昵称": first_name,
                    "用户ID": user_id
                })

            # 创建DataFrame
            df_groups_channels = pd.DataFrame(group_channel_data)
            df_bots = pd.DataFrame(bot_data)

            # 导出到Excel
            file_path = "telegram_info.xlsx"
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df_groups_channels.to_excel(writer, sheet_name="群组和频道", index=False)
                df_bots.to_excel(writer, sheet_name="机器人", index=False)

            # 加载Excel文件并添加筛选功能
            wb = load_workbook(file_path)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # 创建表格
                tab = Table(displayName=f"Table_{sheet_name}", ref=ws.dimensions)
                # 添加样式
                style = TableStyleInfo(
                    name="TableStyleMedium9", showFirstColumn=False,
                    showLastColumn=False, showRowStripes=True, showColumnStripes=True)
                tab.tableStyleInfo = style
                ws.add_table(tab)

                # 自动调整列宽
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            length = get_display_width(cell.value)
                            if length > max_length:
                                max_length = length
                        except:
                            print(traceback.format_exc())
                            pass
                    adjusted_width = (max_length + 4)
                    ws.column_dimensions[column].width = adjusted_width

            wb.save(file_path)
            # 断开客户端
            await client.disconnect()
            return True
        except:
            print('获取失败：{}'.format(traceback.format_exc()))
            client.disconnect()
            return False


class ControllerExport:
    def __init__(self, main_controller, ui):
        self.export_thread = None
        self.main_controller = main_controller
        self.ui = ui
        self.ui.main_button_clicked_signal.connect(self.on_main_button_clicked)

    def on_main_button_clicked(self):
        # 禁用主界面按钮
        self.ui.main_page.main_btn_2.setEnabled(False)
        self.ui.main_page.export_info_label.setText("正在获取数据...")
        self.ui.main_page.export_info_label.setStyleSheet("color: black")

        api_id = self.ui.main_page.id_input.text()
        api_hash = self.ui.main_page.api_hash_input.text()
        phone = self.ui.main_page.phone_input.text()
        proxy_setting = self.main_controller.proxy_settings

        self.export_thread = ExportThread(
            f"telegram_info_{phone}", api_id, api_hash, phone, proxy=proxy_setting, parent=self.ui
        )

        self.export_thread.dialog_signal.connect(self.get_input_from_dialog)
        self.export_thread.finished_signal.connect(self.on_export_finished)
        self.export_thread.start()

    def get_input_from_dialog(self, title, content, show, thread):
        try:
            if show:
                input_text, ok = QInputDialog.getText(self.ui, title, content, QLineEdit.Password)
                if  ok :
                    thread.input_value = input_text
        except Exception as e:
            print(f"对话框处理出错: {e}")
            traceback.print_exc()

    def on_export_finished(self, result):
        # 启用主界面按钮
        if result:
            self.ui.main_page.export_info_label.setText("获取数据成功")
            self.ui.main_page.export_info_label.setStyleSheet("color: green")
            self.ui.main_page.main_btn_2.setEnabled(True)
        else:
            self.ui.main_page.export_info_label.setText("获取数据失败")
            self.ui.main_page.export_info_label.setStyleSheet("color: red")
            self.ui.main_page.main_btn_2.setEnabled(True)
