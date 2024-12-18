import re
import asyncio
import requests
import traceback
import pandas as pd
from telethon import TelegramClient
from app.utils import get_resource_path
from PyQt5.QtCore import QThread, pyqtSignal
from openpyxl.reader.excel import load_workbook
from telethon.errors import SessionPasswordNeededError
from openpyxl.worksheet.table import Table, TableStyleInfo
from PyQt5.QtWidgets import QInputDialog, QMessageBox, QLineEdit


async def get_telegram_info(api_name, api_id, api_hash, phone_number, proxy=None, parent=None):  # 添加parent参数
    def get_display_width(s):
        """
        计算字符串的显示宽度
        中文字符算2个宽度，英文字符算1个宽度
        """
        if not isinstance(s, str):
            s = str(s)
        # 使用正则表达式分辨中文字符和非中文字符
        zh_pattern = re.compile(u'[\u4e00-\u9fa5]')
        width = 0
        for ch in s:
            if zh_pattern.match(ch):
                width += 2
            else:
                width += 1
        return width

    # 创建并连接客户端
    print(api_name, api_id, api_hash, phone_number, proxy)
    client = TelegramClient(api_name, int(api_id), api_hash, proxy=proxy)
    await client.connect()

    if not await client.is_user_authorized():
        # 使用 PyQt 的 QInputDialog 获取验证码
        await client.send_code_request(phone_number)
        text, ok = QInputDialog.getText(parent, '验证码', '请输入验证码:')  # 使用传入的parent
        if not ok:
            QMessageBox.warning(parent, '警告', '未输入验证码，操作取消！')  # 使用传入的parent
            await client.disconnect()
            return
        verification_code = text
        try:
            await client.sign_in(phone_number, verification_code)
        except SessionPasswordNeededError:
            text, ok = QInputDialog.getText(parent, '两步验证', '请输入两步验证密码:', QLineEdit.Password)
            if not ok:
                QMessageBox.warning(parent, '警告', '未输入两步验证密码，操作取消！')
                await client.disconnect()
                return
            password = text
            await client.sign_in(password=password)

    # 获取对话
    groups_and_channels = []
    bots = []

    async for dialog in client.iter_dialogs():
        if dialog.is_group:
            groups_and_channels.append((dialog.entity, '群组', dialog.name))
        elif dialog.is_channel:
            groups_and_channels.append((dialog.entity, '频道', dialog.name))
        elif dialog.entity.bot:
            bots.append((dialog.entity.username, dialog.entity.first_name, dialog.entity.id))

    # 准备导出数据
    group_channel_data = []
    for entity, chat_type, chat_title in groups_and_channels:
        try:
            username = getattr(entity, 'username', None)
            if username:
                invite_link = f"https://t.me/{username}"
            else:
                invite_link = ""
            group_channel_data.append({
                "类型": chat_type,
                "ID": entity.id,
                "名称": chat_title,
                "邀请链接": invite_link
            })
        except Exception:
            print(traceback.format_exc())
            group_channel_data.append({
                "类型": chat_type,
                "ID": entity.id,
                "名称": chat_title,
                "邀请链接": ""
            })

    bot_data = []
    for username, first_name, user_id in bots:
        bot_data.append({
            "用户名": f"@{username}",
            "昵称": first_name,
            "用户ID": user_id
        })

    # 创建DataFrame
    df_groups_channels = pd.DataFrame(group_channel_data)
    df_bots = pd.DataFrame(bot_data)

    # 导出到Excel
    file_path = "telegram_info.xlsx"
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df_groups_channels.to_excel(writer, sheet_name="群组和频道", index=False)
        df_bots.to_excel(writer, sheet_name="机器人", index=False)

    # 加载Excel文件并添加筛选功能
    wb = load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 创建表格
        tab = Table(displayName=f"Table_{sheet_name}", ref=ws.dimensions)
        # 添加样式
        style = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    length = get_display_width(cell.value)
                    if length > max_length:
                        max_length = length
                except:
                    print(traceback.format_exc())
                    pass
            adjusted_width = (max_length + 4)
            ws.column_dimensions[column].width = adjusted_width

    wb.save(file_path)
    # 断开客户端
    await client.disconnect()


def verify_proxy(proxy):
    try:
        proxy_type = 'https'
        if len(proxy) == 3:
            response = requests.get(
                "https://www.google.com",
                proxies={proxy_type: f"{proxy[0].lower()}://{proxy[1]}:{proxy[2]}"},
                timeout=5
            )
        else:
            response = requests.get(
                "https://www.google.com",
                proxies={proxy_type: f"{proxy[0].lower()}://{proxy[3]}:{proxy[4]}@{proxy[1]}:{proxy[2]}"},
                timeout=5
            )
    except:
        print(traceback.format_exc())
        return False
    if response.status_code == 200:
        return True
    else:
        return False


# 创建一个自定义的线程类
class ProxyVerifyThread(QThread):
    # 定义一个信号，用于在验证完成后向主线程传递信息
    finished_signal = pyqtSignal(bool)

    def __init__(self, proxy):
        super().__init__()
        self.proxy = proxy

    def run(self):
        # 执行代理验证逻辑
        verify_success = verify_proxy(self.proxy)
        # 发送信号，传递验证结果
        self.finished_signal.emit(verify_success)


class Controller:
    def __init__(self, ui):
        self.style_path = get_resource_path("style/style.qss")
        self.ui = ui
        self.setup_connections()  # 连接信号和槽函数
        self.ui.load_stylesheet(self.style_path)  # 在这里加载样式表

        self.proxy_settings = None
        self.proxy_thread = None  # 初始化线程对象

        self.is_main_button_running = False

    def setup_connections(self):
        self.ui.main_button_clicked_signal.connect(lambda: asyncio.run(self.on_main_button_clicked()))
        self.ui.save_verify_button_clicked_signal.connect(self.on_save_verify_button_clicked)
        self.ui.set_page_signal.connect(self.on_set_page)

    async def on_main_button_clicked(self):
        # 检查按钮是否已经在运行
        if self.is_main_button_running:
            return
        self.is_main_button_running = True

        # 禁用主界面按钮
        self.ui.main_btn_2.setEnabled(False)
        self.ui.export_info_label.setText("正在获取数据...")
        self.ui.export_info_label.setStyleSheet("color: black")

        api_id = self.ui.id_input.text()
        api_hash = self.ui.api_hash_input.text()
        phone = self.ui.phone_input.text()
        try:
            await get_telegram_info(
                f"telegram_info_{phone}", api_id, api_hash, phone,
                self.proxy_settings, parent=self.ui
            )  # 传入 self.ui 作为 parent
        except:
            print(traceback.format_exc())
            self.ui.export_info_label.setText("获取数据失败")
            self.ui.export_info_label.setStyleSheet("color: red")
            self.ui.main_btn_2.setEnabled(True)
            self.is_main_button_running = False
            return
        # 启用主界面按钮
        self.ui.export_info_label.setText("获取数据成功")
        self.ui.export_info_label.setStyleSheet("color: green")
        self.ui.main_btn_2.setEnabled(True)
        self.is_main_button_running = False
        # print("ID为：" + api_id + ",API Hash为：" + api_hash + ",手机号为：" + phone)

    def on_save_verify_button_clicked(self):
        # 禁用按钮
        self.ui.save_verify_btn.setEnabled(False)
        self.ui.verify_label.setText("正在验证代理...")
        self.ui.verify_label.setStyleSheet("color: black")

        proxy_type = ""
        for radio in self.ui.proxy_type_radio_list:
            if radio.isChecked():
                proxy_type = radio.text()

        proxy_ip = self.ui.proxy_ip_input.text()
        proxy_port = self.ui.proxy_port_input.text()
        username = self.ui.username_input.text()
        password = self.ui.password_input.text()

        # 使用代理访问Google
        if username and password:
            proxy = (proxy_type, proxy_ip, proxy_port, username, password)
        else:
            proxy = (proxy_type, proxy_ip, proxy_port)

        # 创建并启动线程
        self.proxy_thread = ProxyVerifyThread(proxy)
        self.proxy_thread.finished_signal.connect(self.on_verify_finished)
        self.proxy_thread.start()

    def on_verify_finished(self, result):
        if result:
            self.ui.verify_label.setText("配置已保存，代理验证成功")
            self.ui.verify_label.setStyleSheet("color: green")
            # 可以把self.proxy_settings赋值 方便其它函数使用
        else:
            self.ui.verify_label.setText("代理验证失败，请检查配置")
            self.ui.verify_label.setStyleSheet("color: red")
        # 无论如何都需要启用按钮
        self.ui.save_verify_btn.setEnabled(True)

    def on_set_page(self, index):
        self.ui.set_current_page(index)
